#!/usr/bin/env python3
"""
Excel Skill Tracker Migration Script
Migrates skill data from ~381 Excel skill tracker files to Supabase
Updates existing records, doesn't create duplicates
"""

import os
import sys
import argparse
import pandas as pd
from openpyxl import load_workbook
from datetime import datetime
from supabase import create_client, Client
import re
from pathlib import Path

# Load environment variables
def load_env():
    """Load environment variables from .env.local"""
    env_vars = {}
    try:
        with open('.env.local', 'r') as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith('#') and '=' in line:
                    key, value = line.split('=', 1)
                    env_vars[key.strip()] = value.strip()
    except FileNotFoundError:
        print("WARNING: .env.local not found")
    return env_vars

env = load_env()
SUPABASE_URL = env.get('NEXT_PUBLIC_SUPABASE_URL', 'https://jtqlamkrhdfwtmaubfrc.supabase.co')
SUPABASE_KEY = env.get('SUPABASE_SECRET_KEY')  # Use service role key

if not SUPABASE_KEY:
    print("ERROR: SUPABASE_SECRET_KEY not found in .env.local")
    print("Make sure .env.local has SUPABASE_SECRET_KEY from Supabase Dashboard > Settings > API")
    sys.exit(1)

supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

# Skill Name Mapping (Excel → Database)
# Based on Excel skill names from the requirements and previous imports
SKILL_NAME_MAP = {
    # Excel name → Database skill name
    "Asking permission to get in the wat": "Asking permission to get in the water",
    "Asking Permission to get in the water": "Asking permission to get in the water",
    "Asking permission to get in the water": "Asking permission to get in the water",
    "Relaxed Submersion": "Relaxed body position",
    "Starfish float on front and back": "Front float with assistance",
    "Bobbing 5 times": "Blow bubbles",
    "Tuck and stand on back": "Back float with assistance",
    "Kicking on front and back 10 feet": "Kicking with board",
    "Jump or roll in, roll to back and b": "Jump or roll in, roll to back and breath",
    "JUmp or roll in, roll to back and breath": "Jump or roll in, roll to back and breath",
    "Jump or roll in, roll to back and breath": "Jump or roll in, roll to back and breath",
    "Wearing lifejacket and jump in and ": "Wearing lifejacket and jump in and kick on back 10 feet",
    "Wearing lifejacket and jump in and kick on back 10 feet": "Wearing lifejacket and jump in and kick on back 10 feet",
    "Tread water for 10 seconds": "Treading water 10 seconds",
    "Disorientating entries and recover": "Disorientating entries and recover",
    "Reach and throw with assist flotation": "Reach and throw with assist flotation",
    "Roll to the back from the front": "Roll to the back from the front",
    "Swim-roll-swim": "Swim-roll-swim",
    "swim-roll-swim": "Swim-roll-swim",
    "3 stroke- roll and rest": "3 stroke - roll and rest",
    "Beginner stroke on back": "Backstroke 25 yards",  # Approximation
    "Beginner stroke on front- face in": "Front crawl 25 yards",  # Approximation
    "Beginner stroke with direction change": "Front crawl arms",  # Approximation
    "Front and back streamline with kick": "Streamline push-offs",  # Approximation
    "Side breathing position with kick": "Side breathing",  # Approximation
    "Side-roll-side with kick": "Side breathing",  # Approximation
    "Swim underwater 3 feet": "Front crawl arms",  # Approximation
    "Pour water on face and head": "Pour water on face and head",  # Already matches
    "Breath hold and look under water": "Breath hold and look under water",  # Already matches
    "Safely entering and exiting": "Safely entering and exiting",  # Already matches
    "Tuck and stand from front": "Tuck and stand from front",  # Already matches
    "3 stroke - stop drill": "3 stroke - stop drill",  # Already matches
    "3x3 swim drill": "3x3 swim drill",  # Already matches
    "Tread water 40 seconds": "Tread water 40 seconds",  # Already matches
}

# Status mapping from Excel to database
STATUS_MAP = {
    'Not Started': 'not_started',
    'Emerging': 'in_progress',
    'Met': 'mastered',
    'not_started': 'not_started',
    'in_progress': 'in_progress',
    'mastered': 'mastered'
}

def parse_date(date_value):
    """Parse date from Excel cell value"""
    if not date_value or pd.isna(date_value):
        return None

    try:
        # If it's already a datetime object
        if isinstance(date_value, datetime):
            return date_value.date().isoformat()

        # If it's a string
        if isinstance(date_value, str):
            date_str = date_value.strip()
            if not date_str:
                return None

            # Try various date formats
            date_formats = [
                '%Y-%m-%d',
                '%m/%d/%Y',
                '%d/%m/%Y',
                '%Y-%m-%d %H:%M:%S',
                '%m/%d/%Y %H:%M:%S'
            ]

            for fmt in date_formats:
                try:
                    dt = datetime.strptime(date_str, fmt)
                    return dt.date().isoformat()
                except ValueError:
                    continue

            # Try pandas parsing
            try:
                dt = pd.to_datetime(date_str)
                return dt.date().isoformat()
            except:
                return None

        # If it's a number (Excel serial date)
        if isinstance(date_value, (int, float)):
            try:
                dt = pd.Timestamp('1899-12-30') + pd.Timedelta(days=date_value)
                return dt.date().isoformat()
            except:
                return None

    except Exception as e:
        print(f"  Warning: Could not parse date '{date_value}': {e}")

    return None

def extract_swimmer_name(filename):
    """Extract swimmer name from Excel filename"""
    # Pattern: FirstName_LastName_..._Skill_Tracker.xlsx
    filename = Path(filename).stem  # Remove extension
    filename = filename.replace('_Skill_Tracker', '')

    # Remove level indicators
    level_indicators = ['Red_White', 'Red', 'White', 'Yellow', 'Green', 'Blue', 'Master']
    for level in level_indicators:
        filename = filename.replace(f'_{level}', '')

    # Split by underscores and take first two parts as first and last name
    parts = filename.split('_')
    if len(parts) >= 2:
        first_name = parts[0]
        last_name = ' '.join(parts[1:])
        return f"{first_name} {last_name}".strip()
    elif len(parts) == 1:
        return parts[0]
    else:
        return filename

def find_swimmer_in_database(swimmer_name):
    """Find swimmer in database by name"""
    try:
        # Try exact match first
        response = supabase.table('swimmers').select('id, first_name, last_name').execute()

        # Simple fuzzy matching
        swimmer_name_lower = swimmer_name.lower()
        for swimmer in response.data:
            full_name = f"{swimmer['first_name']} {swimmer['last_name']}".lower()
            if swimmer_name_lower == full_name:
                return swimmer['id']

            # Check if names are contained in each other
            if swimmer_name_lower in full_name or full_name in swimmer_name_lower:
                return swimmer['id']

        # Try partial matching
        first_name = swimmer_name.split()[0] if ' ' in swimmer_name else swimmer_name
        for swimmer in response.data:
            if swimmer['first_name'].lower() == first_name.lower():
                return swimmer['id']

    except Exception as e:
        print(f"  Error searching for swimmer '{swimmer_name}': {e}")

    return None

def get_skill_id(skill_name, level_name):
    """Get skill ID from database based on skill name and level"""
    try:
        # First, get the level ID
        level_response = supabase.table('swim_levels').select('id').ilike('name', level_name.lower()).execute()
        if not level_response.data:
            print(f"  Warning: Level '{level_name}' not found in database")
            return None

        level_id = level_response.data[0]['id']

        # Map skill name if needed
        mapped_skill_name = SKILL_NAME_MAP.get(skill_name, skill_name)

        # Try exact match first
        skill_response = supabase.table('skills').select('id').eq('name', mapped_skill_name).eq('level_id', level_id).execute()

        if skill_response.data:
            return skill_response.data[0]['id']

        # Try case-insensitive match
        skill_response = supabase.table('skills').select('id').ilike('name', mapped_skill_name).eq('level_id', level_id).execute()

        if skill_response.data:
            return skill_response.data[0]['id']

        # Try partial match
        all_skills_response = supabase.table('skills').select('id, name').eq('level_id', level_id).execute()
        for skill in all_skills_response.data:
            if mapped_skill_name.lower() in skill['name'].lower() or skill['name'].lower() in mapped_skill_name.lower():
                return skill['id']

        print(f"  Warning: Skill '{mapped_skill_name}' (level: {level_name}) not found in database")

    except Exception as e:
        print(f"  Error finding skill '{skill_name}': {e}")

    return None

def parse_excel_file(file_path):
    """Parse Excel file and extract skill data"""
    print(f"  Parsing: {Path(file_path).name}")

    try:
        wb = load_workbook(file_path, data_only=True)  # data_only=True to get cell values, not formulas
        ws = wb.active

        skills_data = []
        current_section = None

        for row in ws.iter_rows(values_only=True):
            # Skip empty rows
            if not row or all(cell is None for cell in row):
                continue

            # Detect section headers
            if row[0] and isinstance(row[0], str):
                cell_value = str(row[0]).strip()

                if 'Safety Skills' in cell_value:
                    current_section = 'safety'
                    continue
                elif 'Swim Skills' in cell_value:
                    current_section = 'swim'
                    continue
                elif 'I Can Swim Targets' in cell_value:
                    current_section = 'targets'
                    continue
                elif 'Strategies Used' in cell_value:
                    current_section = 'strategies'
                    continue
                elif 'Additional Notes' in cell_value:
                    current_section = 'notes'
                    continue
                elif cell_value in ['Level', 'Target', 'Strategy', 'Date']:
                    continue  # Skip header rows

            # Process skill rows (safety and swim sections)
            if current_section in ['safety', 'swim'] and row[0] and row[1]:
                level = str(row[0]).strip() if row[0] else ''
                skill_name = str(row[1]).strip() if row[1] else ''

                if not level or not skill_name:
                    continue

                status = str(row[2]).strip() if row[2] else 'Not Started'
                date_started = row[3]
                date_met = row[4]
                notes = str(row[5]).strip() if row[5] else ''

                # Map status
                db_status = STATUS_MAP.get(status, 'not_started')

                skills_data.append({
                    'level': level,
                    'skill_name': skill_name,
                    'status': db_status,
                    'date_started': parse_date(date_started),
                    'date_met': parse_date(date_met),
                    'notes': notes,
                    'is_safety_skill': current_section == 'safety'
                })

        return skills_data

    except Exception as e:
        print(f"  Error parsing Excel file: {e}")
        return []

def migrate_excel_file(file_path, test_mode=False):
    """Migrate a single Excel file to database"""
    filename = Path(file_path).name
    swimmer_name = extract_swimmer_name(filename)

    print(f"\nProcessing: {swimmer_name} ({filename})")

    # Find swimmer in database
    swimmer_id = find_swimmer_in_database(swimmer_name)
    if not swimmer_id:
        print(f"  ERROR: Swimmer '{swimmer_name}' not found in database")
        return 0, 0

    # Parse Excel file
    skills_data = parse_excel_file(file_path)
    if not skills_data:
        print(f"  No skill data found in file")
        return 0, 0

    print(f"  Found {len(skills_data)} skill records")

    # Process each skill
    records_to_upsert = []
    skipped_records = 0

    for skill_data in skills_data:
        skill_id = get_skill_id(skill_data['skill_name'], skill_data['level'])
        if not skill_id:
            skipped_records += 1
            continue

        record = {
            'swimmer_id': swimmer_id,
            'skill_id': skill_id,
            'status': skill_data['status'],
            'date_started': skill_data['date_started'],
            'date_met': skill_data['date_met'],
            'instructor_notes': skill_data['notes'] or None,
            'is_safety_skill': skill_data['is_safety_skill']
        }

        records_to_upsert.append(record)

    print(f"  Prepared {len(records_to_upsert)} records for upsert")
    print(f"  Skipped {skipped_records} records (skill not found)")

    if test_mode:
        print(f"  TEST MODE: Would upsert {len(records_to_upsert)} records")
        return len(records_to_upsert), skipped_records

    # Upsert records in batches
    inserted_count = 0
    batch_size = 50

    for i in range(0, len(records_to_upsert), batch_size):
        batch = records_to_upsert[i:i+batch_size]
        try:
            response = supabase.table('swimmer_skills').upsert(
                batch,
                on_conflict='swimmer_id,skill_id'
            ).execute()
            inserted_count += len(batch)
            print(f"    Upserted batch {i//batch_size + 1}: {len(batch)} records")
        except Exception as e:
            print(f"    Error upserting batch {i//batch_size + 1}: {e}")
            # Try individual inserts
            for record in batch:
                try:
                    response = supabase.table('swimmer_skills').upsert(
                        record,
                        on_conflict='swimmer_id,skill_id'
                    ).execute()
                    inserted_count += 1
                except Exception as e2:
                    print(f"      Error upserting individual record: {e2}")

    return inserted_count, skipped_records

def find_excel_files(directory):
    """Find all Excel files in directory"""
    excel_files = []
    for root, dirs, files in os.walk(directory):
        for file in files:
            if file.lower().endswith(('.xlsx', '.xls')):
                excel_files.append(os.path.join(root, file))
    return excel_files

def parse_arguments():
    """Parse command line arguments"""
    parser = argparse.ArgumentParser(description='Migrate Excel skill tracker files to Supabase')
    parser.add_argument('--test', action='store_true', help='Run in test mode (no database changes)')
    parser.add_argument('--directory', type=str, default='.', help='Directory containing Excel files')
    parser.add_argument('--sample', action='store_true', help='Process only first 7 files as sample')
    parser.add_argument('--all', action='store_true', help='Process all files')

    return parser.parse_args()

def main():
    print("=" * 60)
    print("EXCEL SKILL TRACKER MIGRATION")
    print("=" * 60)

    # Parse command line arguments
    args = parse_arguments()

    # Use command line arguments or prompt interactively
    if args.directory and args.directory != '.':
        excel_dir = args.directory
    else:
        excel_dir = input("Enter directory path containing Excel files (or press Enter for current directory): ").strip()
        if not excel_dir:
            excel_dir = "."

    if not os.path.exists(excel_dir):
        print(f"ERROR: Directory '{excel_dir}' does not exist")
        sys.exit(1)

    # Find Excel files
    excel_files = find_excel_files(excel_dir)
    if not excel_files:
        print(f"No Excel files found in '{excel_dir}'")
        sys.exit(1)

    print(f"Found {len(excel_files)} Excel files")

    # Determine test mode
    if args.test:
        test_mode = True
        print("Running in TEST MODE (from command line) - no changes will be made to database")
    else:
        test_mode_input = input("Run in test mode? (y/n, default: y): ").strip().lower()
        test_mode = test_mode_input != 'n'
        if test_mode:
            print("Running in TEST MODE - no changes will be made to database")

    # Determine sample size
    if args.all:
        files_to_process = excel_files
        print("Processing ALL files (from command line)")
    elif args.sample:
        files_to_process = excel_files[:7]
        print("Processing first 7 files as sample (from command line)")
    else:
        sample_input = input(f"Process all {len(excel_files)} files or first 7 as sample? (all/sample, default: sample): ").strip().lower()
        if sample_input == 'all':
            files_to_process = excel_files
        else:
            files_to_process = excel_files[:7]
            print(f"Processing first 7 files as sample")

    # Process files
    total_inserted = 0
    total_skipped = 0
    processed_files = 0

    for file_path in files_to_process:
        inserted, skipped = migrate_excel_file(file_path, test_mode=test_mode)
        total_inserted += inserted
        total_skipped += skipped
        processed_files += 1

    print("\n" + "=" * 60)
    print("MIGRATION SUMMARY")
    print("=" * 60)
    print(f"Processed files: {processed_files}")
    print(f"Records upserted: {total_inserted}")
    print(f"Records skipped: {total_skipped}")

    if test_mode:
        print("\nTEST MODE COMPLETE - No changes were made to database")
        print("To run actual migration, use: python excel_skill_migration.py --directory ./excel-skill-files")
    else:
        print("\nMIGRATION COMPLETE")

    print("=" * 60)

if __name__ == "__main__":
    main()